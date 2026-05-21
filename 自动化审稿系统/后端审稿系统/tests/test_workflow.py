import json
import shutil
import tempfile
import unittest
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook, load_workbook

from journal_automation.config import load_config
from journal_automation.workflow import prepare_reply_package, sync_submissions, update_progress


TOOL_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = Path(__file__).resolve().parents[3]


def build_sample_docx(path: Path, text: str) -> None:
    import zipfile

    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
      <w:body>
        <w:p><w:r><w:t>{text}</w:t></w:r></w:p>
      </w:body>
    </w:document>
    """
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
      <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
      <Default Extension="xml" ContentType="application/xml"/>
      <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
    </Types>
    """
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>
    """
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", rels)
        archive.writestr("word/document.xml", document_xml)


class WorkflowTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = Path(tempfile.mkdtemp(prefix="journal-automation-"))
        (self.temp_dir / "1. 未处理来稿").mkdir()
        (self.temp_dir / "2. 派稿及回复").mkdir()
        (self.temp_dir / "3. 回复作者").mkdir()
        self.workbook_path = self.temp_dir / "台账.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "来稿日期",
            "来稿名称",
            "来稿方向",
            "来稿性质",
            "作者1",
            "校、院、年级、专业方向",
            "作者2",
            "作者院校、年级及专业",
            "一审编辑及审稿情况",
            "二审编辑及审稿情况",
            "再审情况",
            "组内终审情况及拟排期",
            "组内录用时间",
            "作者电话及通讯地址",
            "备注",
        ])
        workbook.save(self.workbook_path)

        self.review_form = self.temp_dir / "审稿评分表.docx"
        shutil.copy2(WORKSPACE_ROOT / "审稿评分表.docx", self.review_form)

        self.law_template = self.temp_dir / "3. 回复作者/中国政法大学研究生学报作者邮件回复模板（法学）.docx"
        self.non_law_template = self.temp_dir / "3. 回复作者/中国政法大学研究生学报作者邮件回复模板（法学以外专业）.docx"
        self.copyright_template = self.temp_dir / "3. 回复作者/附件一、《研究生学报》著作权使用协议模板.docx"
        shutil.copy2(WORKSPACE_ROOT / "3. 回复作者/中国政法大学研究生学报作者邮件回复模板（法学）.docx", self.law_template)
        shutil.copy2(WORKSPACE_ROOT / "3. 回复作者/中国政法大学研究生学报作者邮件回复模板（法学以外专业）.docx", self.non_law_template)
        shutil.copy2(WORKSPACE_ROOT / "3. 回复作者/附件一、《研究生学报》著作权使用协议模板.docx", self.copyright_template)

        self.sample_dir = self.temp_dir / "eml"
        self.sample_dir.mkdir()
        sample_docx = self.temp_dir / "sample.docx"
        build_sample_docx(sample_docx, "论文题目：数据财产民事强制执行路径研究\n作者：张三\n中国政法大学民商法学院2025级硕士研究生")
        msg = EmailMessage()
        msg["Subject"] = "投稿：数据法-数据财产民事强制执行路径研究"
        msg["From"] = "张三 <zhangsan@example.com>"
        msg["To"] = "editor@example.com"
        msg["Date"] = "Tue, 05 May 2026 10:00:00 +0800"
        msg["Message-ID"] = "<sample-1@example.com>"
        msg.set_content("作者：张三\n学校：中国政法大学民商法学院2025级硕士研究生\n联系电话：13800000000\n电子邮箱：zhangsan@example.com\n通讯地址：北京市昌平区")
        msg.add_attachment(
            sample_docx.read_bytes(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename="数据法-数据财产民事强制执行路径研究.docx",
        )
        (self.sample_dir / "sample1.eml").write_bytes(msg.as_bytes())

        config = {
            "root_dir": str(self.temp_dir),
            "mailbox": {
                "imap_host": "imap.126.com",
                "imap_port": 993,
                "smtp_host": "smtp.126.com",
                "smtp_port": 465,
                "username": "editor@example.com",
                "password": "test",
                "folder": "INBOX",
            },
            "runtime": {"state_dir": ".automation"},
            "templates": {
                "review_form": str(self.review_form),
                "law_reply": str(self.law_template),
                "non_law_reply": str(self.non_law_template),
                "copyright_agreement": str(self.copyright_template),
            },
            "workbook": str(self.workbook_path),
            "non_law_subjects": ["管理学"],
            "subject_aliases": {"数据法": ["算法", "人工智能"], "民事诉讼法": ["民诉法"]},
            "editors_by_subject": {"数据法": ["吴秋雨", "韩佩芝"]},
            "preferred_extensions": [".docx", ".doc"],
        }
        self.config_path = self.temp_dir / "automation.config.json"
        self.config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        self.config = load_config(self.config_path)

    def tearDown(self):
        shutil.rmtree(self.temp_dir)

    def test_sync_submissions_from_eml(self):
        result = sync_submissions(self.config, eml_dir=self.sample_dir)
        self.assertEqual(result["created"], 1)
        archived = list((self.temp_dir / "1. 未处理来稿").glob("数据法-数据财产民事强制执行路径研究*.docx"))
        self.assertTrue(archived)
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["B2"].value, "数据财产民事强制执行路径研究")

    def test_generate_materials_and_update_progress(self):
        sync_submissions(self.config, eml_dir=self.sample_dir)
        result = prepare_reply_package(
            self.config,
            record_id=1,
            result_label="一审返修",
            editor_name="张芮铭",
            folder_date="20260505",
        )
        self.assertTrue(Path(result["draft_path"]).exists())
        self.assertTrue(Path(result["stage_folder"]).exists())
        note = update_progress(
            self.config,
            record_id=1,
            result_label="一审返修",
            editor_name="张芮铭",
            note_date="2026.05.05",
        )
        self.assertIn("2026.05.05", note)
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        sheet = workbook.active
        self.assertIn("张芮铭", sheet["I2"].value)


if __name__ == "__main__":
    unittest.main()
