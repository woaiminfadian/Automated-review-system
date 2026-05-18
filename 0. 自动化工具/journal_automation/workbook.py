from __future__ import annotations

from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

from .models import SubmissionRecord


# 总表列名映射。
# 注意：实际总表第一列表头是投稿作者姓名（如"郭怡杉"），但该列实际内容是来稿日期。
# 因不是标准列名，写入时用列位置定位（见 append_submission）。
HEADER_MAP = {
    "来稿名称": "title",
    "来稿方向": "discipline",
    "来稿性质": "discipline",
    "作者1": "author1",
    "校、院、年级、专业方向": "author_info",
    "作者2": "author2",
    "作者院校、年级及专业": "author2_info",
    "作者电话及通讯地址": "contact_info",
    "备注": "remarks",
}


def append_submission(workbook_path: Path, record: SubmissionRecord) -> int:
    workbook = load_workbook(workbook_path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    next_row = sheet.max_row + 1

    # 第一列是日期列，不论表头是"来稿日期"还是责任编辑姓名，都按列位置写入
    sheet.cell(next_row, 1).value = record.sent_at or ""

    values: Dict[str, Optional[str]] = {
        "title": record.title,
        "discipline": record.discipline,
        "author1": record.authors[0] if record.authors else "",
        "author2": record.authors[1] if len(record.authors) > 1 else "",
        "author_info": record.author_info,
        "author2_info": "",
        "contact_info": record.contact_info,
        "remarks": build_remark(record),
    }
    for header, key in HEADER_MAP.items():
        column = headers.get(header)
        if column:
            sheet.cell(next_row, column).value = values.get(key, "")
    workbook.save(workbook_path)
    return next_row


def find_duplicate_title(workbook_path: Path, title: str) -> bool:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1) if cell.value}
    title_col = headers.get("来稿名称")
    if not title_col:
        return False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cell = row[title_col - 1]
        if isinstance(cell, str) and cell.strip() == title.strip():
            return True
    return False


def append_progress_note(workbook_path: Path, row_number: int, column_header: str, note: str) -> None:
    workbook = load_workbook(workbook_path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    column = headers[column_header]
    current = sheet.cell(row_number, column).value or ""
    current = str(current).strip()
    if current:
        sheet.cell(row_number, column).value = current + "\n" + note
    else:
        sheet.cell(row_number, column).value = note
    workbook.save(workbook_path)


def build_remark(record: SubmissionRecord) -> str:
    items = []
    if record.needs_manual_review:
        items.append("待人工确认")
    if record.needs_anonymization_check:
        items.append("疑似未匿名")
    if record.duplicate_warning:
        items.append("可能重复投稿")
    return "；".join(items)

